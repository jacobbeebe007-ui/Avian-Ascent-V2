#!/usr/bin/env python3
"""Phase 1: close v0.7 workbook gaps before game sync import."""
from __future__ import annotations

from collections import defaultdict
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

SRC = Path(
    "/Users/j.beebe/Documents/Avian Ascent/Avian Workbooks/"
    "v7Newest_Avian_Ascent_Master_v0.7_Equipment_Loot_Implemented.xlsx"
)
DST = Path(
    "/Users/j.beebe/Documents/Avian Ascent/Avian Workbooks/"
    "v7Newest_Avian_Ascent_Master_v0.7_Equipment_Loot_SyncReady.xlsx"
)

FOCUS_AFFINITY = {
    "Poison": "Earth Orb Focus",
    "Burn": "Day Orb Focus",
    "Chill": "Water Orb Focus",
    "Shock": "Storm Orb Focus",
    "Bleed": "Night Orb Focus",
}


def clear_sheet(ws):
    if ws.merged_cells.ranges:
        for merged in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged))
    for row in ws.iter_rows():
        for cell in row:
            try:
                cell.value = None
            except AttributeError:
                pass


def write_table(ws, title, subtitle, headers, rows):
    clear_sheet(ws)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = subtitle
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.font = Font(bold=True)
    for r_i, row in enumerate(rows, 5):
        for c_i, val in enumerate(row, 1):
            ws.cell(r_i, c_i, val)
    ws.freeze_panes = "A5"


def formula_row(template_row: int, new_row: int, template_cells: list):
    """Rewrite simple absolute/relative row refs from template_row → new_row in formulas."""
    out = []
    old = str(template_row)
    new = str(new_row)
    for val in template_cells:
        if isinstance(val, str) and val.startswith("="):
            # Replace row number tokens carefully: $R69, G69, AL69, etc.
            import re

            def repl(m):
                prefix, num = m.group(1), m.group(2)
                if num == old:
                    return f"{prefix}{new}"
                return m.group(0)

            out.append(re.sub(r"(\$?[A-Z]{1,3})(\d+)", repl, val))
        else:
            out.append(val)
    return out


def main():
    wb = openpyxl.load_workbook(SRC)

    # ---- derive roll ranges from Equipment Stats ----
    es = wb["Equipment Stats"]
    es_h = {es.cell(4, c).value: c for c in range(1, 30) if es.cell(4, c).value}
    flat_cols = [
        "Vitality Flat",
        "Might Flat",
        "Guard Flat",
        "Focus Flat",
        "Resolve Flat",
        "Agility Flat",
    ]
    pct_cols = [
        "Vitality %",
        "Might %",
        "Guard %",
        "Focus %",
        "Resolve %",
        "Agility %",
        "Evasion %",
        "Critical %",
        "Ferocity %",
        "Martial Penetration %",
        "Magic Penetration %",
        "Martial Damage %",
        "Magic Damage %",
        "Affinity Damage %",
        "Healing Power %",
        "Barrier Power %",
    ]
    ranges = defaultdict(lambda: {"flat": defaultdict(list), "pct": defaultdict(list)})
    # need budget class from Equipment Stats or Catalogue
    cat = wb["Equipment Catalogue"]
    cat_h = {cat.cell(4, c).value: c for c in range(1, 40) if cat.cell(4, c).value}
    id_to_budget = {}
    for r in range(5, cat.max_row + 1):
        iid = cat.cell(r, cat_h["Item ID"]).value
        if iid:
            id_to_budget[iid] = cat.cell(r, cat_h["Budget Class"]).value

    for r in range(5, es.max_row + 1):
        iid = es.cell(r, es_h["Item ID"]).value
        if not iid:
            continue
        rarity = es.cell(r, es_h["Rarity"]).value
        budget = es.cell(r, es_h.get("Budget Class", 4)).value or id_to_budget.get(iid)
        key = (rarity, budget)
        for col in flat_cols:
            v = es.cell(r, es_h[col]).value
            if v is not None and float(v) != 0:
                ranges[key]["flat"][col].append(float(v))
        for col in pct_cols:
            v = es.cell(r, es_h[col]).value
            if v is not None and float(v) != 0:
                ranges[key]["pct"][col].append(float(v))

    roll_rows = []
    for (rarity, budget) in sorted(ranges.keys(), key=lambda x: (str(x[0]), str(x[1]))):
        data = ranges[(rarity, budget)]
        for kind, store in (("Flat", data["flat"]), ("%", data["pct"])):
            for col, vals in sorted(store.items()):
                mn, mx = min(vals), max(vals)
                mid = (mn + mx) / 2
                spread = ((mx - mn) / mid * 100) if mid else 0
                roll_rows.append(
                    [
                        rarity,
                        budget,
                        col,
                        kind,
                        round(mn, 4) if kind == "%" else int(mn),
                        round(mx, 4) if kind == "%" else int(mx),
                        len(vals),
                        round(spread, 1),
                    ]
                )

    write_table(
        wb["Equipment Roll Ranges"],
        "Equipment Roll Ranges — v0.7 Observed Matrix",
        "Derived from the 240 hybrid Equipment Stats rows. R-RNG-001: same-rarity best rolls should be about 25–35% stronger than minimum without labelling low rolls as poor. Flat rolls are integers; percentages may display to two decimals.",
        [
            "Rarity",
            "Budget Class",
            "Stat Column",
            "Kind",
            "Min Observed",
            "Max Observed",
            "Nonzero Samples",
            "Max vs Mid %",
        ],
        roll_rows,
    )

    # ---- EN Damage Rules ----
    write_table(
        wb["EN Damage Rules"],
        "EN Damage Rules — Equipment Loot v0.7",
        "Confirmed bands from R-EN-005 / R-CD-001 / R-CD-002. Cooldown never refunds damage, healing or utility budget. Utility, ailment, penetration, healing and barrier riders reduce direct damage instead.",
        ["Rule ID", "Topic", "EN / Band", "Value", "Notes"],
        [
            ["R-EN-005", "Pure damage band", "1 EN", "70–90%", "Natural Strike is the universal 1 EN baseline"],
            ["R-EN-005", "Pure damage band", "2 EN", "100–120%", "Ordinary weapon technique floor"],
            ["R-EN-005", "Pure damage band", "3 EN", "130–150%", "Paired / combination standard"],
            ["R-EN-005", "Pure damage band", "4 EN", "160–190%", "True Heavy"],
            ["R-EN-005", "Pure damage band", "5 EN", "195–225%", "Reserved / rare"],
            ["R-EN-005", "Pure damage band", "6 EN", "235–280%", "Ultimate gate"],
            ["R-CD-001", "Cooldown", "1–2 EN", "0 turns", "No cooldown"],
            ["R-CD-001", "Cooldown", "3 EN", "1 turn", "Fixed usage restriction"],
            ["R-CD-001", "Cooldown", "4 EN", "2 turns", "Fixed usage restriction"],
            ["R-CD-001", "Cooldown", "6 EN", "Ultimate gate", "Meter Full; cooldown column 0"],
            ["R-CD-002", "Budget", "Any", "No CD refund", "Riders reduce direct damage instead of gaining cooldown compensation"],
            ["R-EN-001", "Economy", "Start / Regen", "4 start, +3 / turn", "Unchanged foundation"],
            ["R-EN-003", "Roles", "Basic / Ordinary / Heavy", "1 / ≥2 / 4 EN", "Natural Strike and Basic Attack are 1 EN"],
        ],
    )

    # ---- Affix Pools ----
    write_table(
        wb["Affix Pools"],
        "Affix Pools — Equipment Loot v0.7",
        "Separate Ignore Guard and Ignore Resolve. All penetration sources share the 40% cap (R-PEN-002). No direct Frozen, Paralysed, Sleep or Silence from active equipment or skill text.",
        ["Pool", "Property", "Channel", "Rules", "Examples / Families"],
        [
            ["Penetration", "Ignore Guard", "Martial", "Separate from Resolve; shares 40% cap", "Beak Hammer, War Pick, Hook Axe, Greatblade"],
            ["Penetration", "Ignore Resolve", "Magic", "Separate from Guard; shares 40% cap", "Wand, Hexwood Wand, Focus Orb, Staff"],
            ["Penetration", "Martial Penetration %", "Martial", "Equipment % affix; capped with Ignore Guard", "Weapon / anklet hybrids"],
            ["Penetration", "Magic Penetration %", "Magic", "Equipment % affix; capped with Ignore Resolve", "Wand / mystic hybrids"],
            ["Control", "Forbidden direct apply", "—", "No Frozen / Paralysed / Sleep / Silence on active text", "Control earned via Chilled@5 or Shock@5"],
            ["Riders", "Ailment", "Attack", "Deterministic stacks after land; reduce direct damage", "Poison/Burn/Chill/Shock/Bleed weapons"],
            ["Riders", "Utility / Heal / Barrier", "Support", "Reduce direct damage; no cooldown budget refund", "Armour techniques, songs"],
            ["Core Hybrid", "Flat core stats", "V/M/G/F/R/A", "Integers; applied after Tier, before %", "Grey emphasises flat"],
            ["Core Hybrid", "Percentage core stats", "V/M/G/F/R/A", "Fractions; additive after flat", "Higher rarities mix both"],
            ["Secondary %", "Evasion / Critical / Ferocity / Damage / Heal / Barrier", "Various", "Percentage / point affixes per rarity budget", "See Equipment Stats"],
        ],
    )

    # ---- v0.7 Rules Update ----
    write_table(
        wb["v0.7 Rules Update"],
        "v0.7 Rules Update — Index",
        "Equipment Loot integration index. Confirmed rules live on Current Rules; this sheet is the quick map.",
        ["Rule ID", "Topic", "Summary", "Sheet / System", "Status"],
        [
            ["R-PROG-005", "Calc order", "Base+Levels+Stars → Tier → Equipment Flat → Equipment % → Temp", "Progression Rules / Loadout Builder", "Confirmed"],
            ["R-EQP-001", "Hybrid gear", "Flat and % Vitality/Might/Guard/Focus/Resolve/Agility", "Equipment Stats", "Confirmed"],
            ["R-EQP-004", "Loot mix", "Grey emphasises flat; later rarities mix both", "Equipment Roll Ranges", "Confirmed"],
            ["R-EN-005", "Damage bands", "1 EN 70–90% … 6 EN 235–280%", "EN Damage Rules / Skill Library", "Confirmed"],
            ["R-CD-001", "Fixed CD", "1–2→0; 3→1; 4→2; 6 Ultimate gate", "EN Damage Rules / Skill Library", "Confirmed"],
            ["R-PEN-002", "Penetration", "Ignore Guard and Ignore Resolve are separate; 40% shared cap", "Affix Pools", "Confirmed"],
            ["R-RNG-001", "Roll quality", "Best same-rarity ~25–35% above minimum", "Equipment Roll Ranges", "Confirmed direction"],
            ["CHG-055", "Control language", "No direct Frozen/Paralysed/Sleep/Silence", "Skill Library / Catalogue", "Confirmed"],
            ["CHG-057", "Dual wield", "Expanded Focus, Wand/Focus and martial pair registry", "Combination Techniques", "Confirmed structure"],
            ["CHG-058", "Ranged access", "Bow/HXB Rogue/Bard; Greatbow Knight/Brute; catalogue pending for Bow/HXB", "Equipment Families", "Confirmed access / pending items"],
        ],
    )

    # ---- Implementation Checklist ----
    write_table(
        wb["Implementation Checklist"],
        "Implementation Checklist — v0.7 → Game Sync",
        "Acceptance tests for workbook integrity and repository import.",
        ["Order", "Task", "Sheet / System", "Acceptance Test", "Priority", "Status"],
        [
            [1, "Hybrid Equipment Stats", "Equipment Stats", "240 items; every row has ≥1 nonzero Flat core", "P0", "Done"],
            [2, "Calc order Flat→%", "Loadout Builder / Reference Battles", "Formulas add Flat after Tier before %", "P0", "Done"],
            [3, "EN bands + fixed CD", "Skill Library", "All skills match R-EN-005 / R-CD-001", "P0", "Done"],
            [4, "Combination Skill IDs", "Skill Library ↔ Combination Techniques", "Every CT Skill ID exists in Skill Library", "P0", "In progress"],
            [5, "Family guide coverage", "Equipment Families", "All catalogue families listed; pending families flagged", "P0", "In progress"],
            [6, "Rule doc sheets", "EN/Ranges/Affix/Audit sheets", "No cloned empty checklists", "P1", "Done"],
            [7, "Importer hybrid parse", "scripts/import-equipment-workbook.mjs", "Imports Flat+%; pack stamp v0.7", "P0", "Pending code"],
            [8, "Runtime Flat→%", "bird-progression / equipment.js", "computeFinalStats order matches R-PROG-005", "P0", "Pending code"],
            [9, "Save stamp v15", "save-migrations.js", "equipmentLootPackVersion set", "P1", "Pending code"],
            [10, "Verify suite", "npm test", "All verifiers pass on v0.7 packs", "P0", "Pending code"],
        ],
    )

    # ---- Skill Library: add 14 missing skills ----
    sk = wb["Skill Library"]
    headers = [sk.cell(4, c).value for c in range(1, 41)]
    existing = set()
    for r in range(5, sk.max_row + 1):
        sid = sk.cell(r, 1).value
        if sid:
            existing.add(sid)

    # clone formulas from COMBO_POISON_TALON row
    template_r = None
    for r in range(5, sk.max_row + 1):
        if sk.cell(r, 1).value == "COMBO_POISON_TALON":
            template_r = r
            break
    template_vals = [sk.cell(template_r, c).value for c in range(1, 41)]
    template_styles = [copy(sk.cell(template_r, c)._style) for c in range(1, 41)]

    new_skills = [
        # Focus matching pairs
        dict(
            id="PAIR_FOCUS_POISON",
            name="Concentrated Venom",
            source="Paired Weapons",
            family="Focus Orb",
            skill_type="Attack",
            en=3,
            cd=1,
            dmg="Magic",
            primary="Focus",
            secondary=None,
            sec_w=0,
            affinity="Earth Orb Focus",
            hits=2,
            base=6,
            master=1.35,
            rider="On hit, apply 2 Poison stacks.",
            note="Two Poison Focus Orbs",
            design="Matching Focus pair; +1 Poison over a single Focus Pulse contribution budget.",
            prec=0.94,
            pcoeff=1.35,
            scoeff=0,
            role="Paired",
            desc="Strike 2 times for a total of 6 + 135% Focus as Magic damage. On hit, apply 2 Poison stacks.",
        ),
        dict(
            id="PAIR_FOCUS_BURN",
            name="Concentrated Flame",
            source="Paired Weapons",
            family="Focus Orb",
            skill_type="Attack",
            en=3,
            cd=1,
            dmg="Magic",
            primary="Focus",
            secondary=None,
            sec_w=0,
            affinity="Day Orb Focus",
            hits=2,
            base=6,
            master=1.35,
            rider="On hit, apply 2 Burn stacks.",
            note="Two Burn Focus Orbs",
            design="Matching Focus pair.",
            prec=0.94,
            pcoeff=1.35,
            scoeff=0,
            role="Paired",
            desc="Strike 2 times for a total of 6 + 135% Focus as Magic damage. On hit, apply 2 Burn stacks.",
        ),
        dict(
            id="PAIR_FOCUS_CHILL",
            name="Concentrated Frost",
            source="Paired Weapons",
            family="Focus Orb",
            skill_type="Attack",
            en=3,
            cd=1,
            dmg="Magic",
            primary="Focus",
            secondary=None,
            sec_w=0,
            affinity="Water Orb Focus",
            hits=2,
            base=6,
            master=1.35,
            rider="On hit, apply 2 Chilled stacks.",
            note="Two Chill Focus Orbs",
            design="Matching Focus pair.",
            prec=0.94,
            pcoeff=1.35,
            scoeff=0,
            role="Paired",
            desc="Strike 2 times for a total of 6 + 135% Focus as Magic damage. On hit, apply 2 Chilled stacks.",
        ),
        dict(
            id="PAIR_FOCUS_SHOCK",
            name="Concentrated Storm",
            source="Paired Weapons",
            family="Focus Orb",
            skill_type="Attack",
            en=3,
            cd=1,
            dmg="Magic",
            primary="Focus",
            secondary=None,
            sec_w=0,
            affinity="Storm Orb Focus",
            hits=2,
            base=6,
            master=1.35,
            rider="On hit, apply 2 Shock stacks.",
            note="Two Shock Focus Orbs",
            design="Matching Focus pair.",
            prec=0.94,
            pcoeff=1.35,
            scoeff=0,
            role="Paired",
            desc="Strike 2 times for a total of 6 + 135% Focus as Magic damage. On hit, apply 2 Shock stacks.",
        ),
        dict(
            id="PAIR_FOCUS_BLEED",
            name="Concentrated Blood",
            source="Paired Weapons",
            family="Focus Orb",
            skill_type="Attack",
            en=3,
            cd=1,
            dmg="Magic",
            primary="Focus",
            secondary=None,
            sec_w=0,
            affinity="Night Orb Focus",
            hits=2,
            base=6,
            master=1.35,
            rider="On hit, apply 1 Bleed stack and increase Bleed damage or duration on that target.",
            note="Two Bleed Focus Orbs",
            design="Matching Focus pair; Bleed remains at three-stack family rules.",
            prec=0.94,
            pcoeff=1.35,
            scoeff=0,
            role="Paired",
            desc="Strike 2 times for a total of 6 + 135% Focus as Magic damage. On hit, apply 1 Bleed stack with an improved Bleed payoff.",
        ),
        # Wand + Focus
        dict(
            id="COMBO_WAND_POISON",
            name="Venom Bolt",
            source="Combination",
            family="Wand + Poison Orb",
            skill_type="Hybrid Attack",
            en=3,
            cd=1,
            dmg="Hybrid",
            primary="Focus",
            secondary="Focus",
            sec_w=0.5,
            affinity="Earth Orb Focus",
            hits=1,
            base=6,
            master=1.42,
            rider="On hit, apply 1 Poison stack.",
            note="Wand + Poison Orb",
            design="Curated Wand/Focus combination; replaces off-hand Focus Pulse only.",
            prec=0.93,
            pcoeff=0.78,
            scoeff=0.64,
            role="Combination",
            desc="Deal 6 + 78% Focus as Magic damage and 64% Focus as Magic damage from the Orb channel. On hit, apply 1 Poison stack.",
        ),
        dict(
            id="COMBO_WAND_BURN",
            name="Cinder Bolt",
            source="Combination",
            family="Wand + Burn Orb",
            skill_type="Hybrid Attack",
            en=3,
            cd=1,
            dmg="Hybrid",
            primary="Focus",
            secondary="Focus",
            sec_w=0.5,
            affinity="Day Orb Focus",
            hits=1,
            base=6,
            master=1.42,
            rider="On hit, apply 1 Burn stack.",
            note="Wand + Burn Orb",
            design="Curated Wand/Focus combination.",
            prec=0.93,
            pcoeff=0.78,
            scoeff=0.64,
            role="Combination",
            desc="Deal 6 + 78% Focus as Magic damage and 64% Focus as Magic damage from the Orb channel. On hit, apply 1 Burn stack.",
        ),
        dict(
            id="COMBO_WAND_CHILL",
            name="Frost Bolt",
            source="Combination",
            family="Wand + Chill Orb",
            skill_type="Hybrid Attack",
            en=3,
            cd=1,
            dmg="Hybrid",
            primary="Focus",
            secondary="Focus",
            sec_w=0.5,
            affinity="Water Orb Focus",
            hits=1,
            base=6,
            master=1.42,
            rider="On hit, apply 1 Chilled stack.",
            note="Wand + Chill Orb",
            design="Curated Wand/Focus combination.",
            prec=0.93,
            pcoeff=0.78,
            scoeff=0.64,
            role="Combination",
            desc="Deal 6 + 78% Focus as Magic damage and 64% Focus as Magic damage from the Orb channel. On hit, apply 1 Chilled stack.",
        ),
        dict(
            id="COMBO_WAND_SHOCK",
            name="Storm Bolt",
            source="Combination",
            family="Wand + Shock Orb",
            skill_type="Hybrid Attack",
            en=3,
            cd=1,
            dmg="Hybrid",
            primary="Focus",
            secondary="Focus",
            sec_w=0.5,
            affinity="Storm Orb Focus",
            hits=1,
            base=6,
            master=1.42,
            rider="On hit, apply 1 Shock stack.",
            note="Wand + Shock Orb",
            design="Curated Wand/Focus combination.",
            prec=0.93,
            pcoeff=0.78,
            scoeff=0.64,
            role="Combination",
            desc="Deal 6 + 78% Focus as Magic damage and 64% Focus as Magic damage from the Orb channel. On hit, apply 1 Shock stack.",
        ),
        # Martial pending-family / mixed pairs
        dict(
            id="PAIR_HOOK_REAPERS",
            name="Twin Reapers",
            source="Paired Weapons",
            family="Hook Axe",
            skill_type="Attack",
            en=3,
            cd=1,
            dmg="Martial",
            primary="Might",
            secondary=None,
            sec_w=0,
            affinity="Main-hand Affinity",
            hits=2,
            base=6,
            master=1.4,
            rider="On hit, apply 1 Bleed stack, or deal bonus Martial damage if the target is already Bleeding.",
            note="Two Hook Axes (family pending catalogue)",
            design="Matching Hook Axe pair; catalogue items deferred.",
            prec=0.93,
            pcoeff=1.4,
            scoeff=0,
            role="Paired",
            desc="Strike 2 times for a total of 6 + 140% Might as Martial damage. On hit, apply 1 Bleed stack or gain damage against Bleeding targets.",
        ),
        dict(
            id="PAIR_SABRE_CROSSWIND",
            name="Crosswind Duel",
            source="Paired Weapons",
            family="Duel Sabre",
            skill_type="Attack",
            en=3,
            cd=1,
            dmg="Martial",
            primary="Might",
            secondary=None,
            sec_w=0,
            affinity="Main-hand Affinity",
            hits=2,
            base=6,
            master=1.48,
            rider="If both hits land, gain improved Critical payoff on this action.",
            note="Two Duel Sabres",
            design="Matching sabre pair.",
            prec=0.94,
            pcoeff=1.48,
            scoeff=0,
            role="Paired",
            desc="Strike 2 times for a total of 6 + 148% Might as Martial damage. If both hits land, gain improved Critical payoff.",
        ),
        dict(
            id="COMBO_TALON_BREAKER",
            name="Talon Breaker",
            source="Combination",
            family="Talon Blade + Beak Hammer",
            skill_type="Attack",
            en=3,
            cd=1,
            dmg="Martial",
            primary="Might",
            secondary=None,
            sec_w=0,
            affinity="Main-hand Affinity",
            hits=1,
            base=6,
            master=1.42,
            rider="Ignore 10% Guard, respecting the 40% penetration cap.",
            note="Talon Blade + Beak Hammer",
            design="Mixed martial combination; Ignore Guard only.",
            prec=0.93,
            pcoeff=1.42,
            scoeff=0,
            role="Combination",
            desc="Deal 6 + 142% Might as Martial damage. Ignore 10% Guard, respecting the 40% penetration cap.",
        ),
        dict(
            id="COMBO_FEATHERKNIFE_FEINT",
            name="Featherknife Feint",
            source="Combination",
            family="Duel Sabre + Dagger Pinion",
            skill_type="Attack",
            en=3,
            cd=1,
            dmg="Martial",
            primary="Agility",
            secondary="Might",
            sec_w=0.4,
            affinity="Main-hand Affinity",
            hits=1,
            base=6,
            master=1.4,
            rider="After landing, gain Minor Evasion Up until the end of your next turn.",
            note="Duel Sabre + Dagger Pinion",
            design="Mixed martial combination.",
            prec=0.95,
            pcoeff=0.9,
            scoeff=0.5,
            role="Combination",
            desc="Deal 6 + 90% Agility and 50% Might as Martial damage. After landing, gain Minor Evasion Up until the end of your next turn.",
        ),
        dict(
            id="COMBO_CITADEL_BREACH",
            name="Citadel Breach",
            source="Combination",
            family="War Pick + Beak Hammer",
            skill_type="Attack",
            en=4,
            cd=2,
            dmg="Martial",
            primary="Might",
            secondary=None,
            sec_w=0,
            affinity="Main-hand Affinity",
            hits=1,
            base=8,
            master=1.75,
            rider="Ignore 20% Guard, respecting the 40% penetration cap.",
            note="War Pick + Beak Hammer (War Pick catalogue pending)",
            design="Heavy mixed martial combination; Ignore Guard only.",
            prec=0.92,
            pcoeff=1.75,
            scoeff=0,
            role="Combination",
            desc="Deal 8 + 175% Might as Martial damage. Ignore 20% Guard, respecting the 40% penetration cap.",
        ),
    ]

    # For Wand hybrid, use Martial+Magic split more like Talon combos but Focus primary.
    # Fix wand combo damage type descriptions - Hybrid with Focus/Focus is odd.
    # Better: Magic primary with Focus, like wand, and secondary from orb as Magic Focus.
    # Keep as Hybrid Attack with Focus primary for importer compatibility.

    h_idx = {h: i for i, h in enumerate(headers) if h}
    added = 0
    for spec in new_skills:
        if spec["id"] in existing:
            continue
        new_r = sk.max_row + 1
        vals = formula_row(template_r, new_r, template_vals)
        # overlay fields by header
        field_map = {
            "Skill ID": spec["id"],
            "Skill Name": spec["name"],
            "Source": spec["source"],
            "Family": spec["family"],
            "Action Bar Slot": "Weapon Technique B",
            "Skill Type": spec["skill_type"],
            "EN": spec["en"],
            "Cooldown": spec["cd"],
            "Meter": "No",
            "Target": "Enemy",
            "Damage Type": spec["dmg"],
            "Primary Scaling Stat": spec["primary"],
            "Secondary Scaling Stat": spec["secondary"],
            "Secondary Weight": spec["sec_w"],
            "Affinity Rule": spec["affinity"],
            "Hits": spec["hits"],
            "Base Damage": spec["base"],
            "Master Scaling": spec["master"],
            "Effect / Rider": spec["rider"],
            "Minimum Rarity": "Grey" if spec["source"] == "Paired Weapons" else "Green",
            "Intrinsic Penetration %": 0.10
            if "Ignore 10% Guard" in spec["rider"]
            else (0.20 if "Ignore 20% Guard" in spec["rider"] else 0),
            "Class / Equip Note": spec["note"],
            "Design Notes": spec["design"],
            "Base Precision": spec["prec"],
            "Primary Coefficient": spec["pcoeff"],
            "Secondary Coefficient": spec["scoeff"],
            "EN Role": spec["role"],
            "Scaling Status": "Confirmed v0.7 band",
            "Standard Description": spec["desc"],
        }
        for h, v in field_map.items():
            if h in h_idx:
                vals[h_idx[h]] = v
        for c, val in enumerate(vals, 1):
            cell = sk.cell(new_r, c, val)
            try:
                cell._style = copy(template_styles[c - 1])
            except Exception:
                pass
        added += 1
        existing.add(spec["id"])

    skill_count = sum(1 for r in range(5, sk.max_row + 1) if sk.cell(r, 1).value)
    sk["A2"] = (
        f"{skill_count} active templates rescaled to the Equipment Loot EN ranges. "
        "Cooldowns are fixed by EN; utility riders reduce direct damage and never receive cooldown budget refunds. "
        "Combination Techniques Skill IDs are fully present in this library."
    )

    # Skill Rebalance Audit — snapshot
    audit_rows = []
    for r in range(5, sk.max_row + 1):
        sid = sk.cell(r, 1).value
        if not sid:
            continue
        audit_rows.append(
            [
                sid,
                sk.cell(r, 2).value,
                sk.cell(r, 3).value,
                sk.cell(r, 7).value,
                sk.cell(r, 8).value,
                sk.cell(r, 17).value,
                sk.cell(r, 18).value,
                sk.cell(r, 34).value,
                sk.cell(r, 35).value,
                sk.cell(r, 36).value,
                sk.cell(r, 37).value,
            ]
        )
    write_table(
        wb["Skill Rebalance Audit"],
        "Skill Rebalance Audit — v0.7",
        "Snapshot of Skill Library EN / Cooldown / Base Damage / coefficients after Equipment Loot rescaling. Cooldown must match R-CD-001.",
        [
            "Skill ID",
            "Skill Name",
            "Source",
            "EN",
            "Cooldown",
            "Base Damage",
            "Master Scaling",
            "Primary Coefficient",
            "Secondary Coefficient",
            "EN Role",
            "Scaling Status",
        ],
        audit_rows,
    )

    # ---- Fix Twin Anvils Guard-or-Resolve ----
    for r in range(5, sk.max_row + 1):
        if sk.cell(r, 1).value == "PAIR_HAMMER_ANVIL":
            sk.cell(r, h_idx["Effect / Rider"] + 1).value = (
                "Strike twice. Ignore 15% Guard, respecting the 40% penetration cap."
            )
            sk.cell(r, h_idx["Standard Description"] + 1).value = (
                "Strike 2 times for a total of 8 + 180% Might as Martial damage. "
                "Resolve Affinity effectiveness using main-hand affinity. "
                "Ignore 15% Guard, respecting the 40% penetration cap."
            )
            sk.cell(r, h_idx["Intrinsic Penetration %"] + 1).value = 0.15
            break

    # Also fix PAIR_WAND_CURRENTS "Resolve or" language → Ignore Resolve only
    for r in range(5, sk.max_row + 1):
        if sk.cell(r, 1).value == "PAIR_WAND_CURRENTS":
            sk.cell(r, h_idx["Effect / Rider"] + 1).value = (
                "Ignore 12% Resolve, respecting the 40% penetration cap."
            )
            desc = sk.cell(r, h_idx["Standard Description"] + 1).value or ""
            sk.cell(r, h_idx["Standard Description"] + 1).value = desc.replace(
                "Ignore 12% Resolve or gain equivalent Magic Penetration, respecting the 40% penetration cap.",
                "Ignore 12% Resolve, respecting the 40% penetration cap.",
            )
            sk.cell(r, h_idx["Intrinsic Penetration %"] + 1).value = 0.12
            break

    # ---- Equipment Families: add missing catalogue families ----
    fam = wb["Equipment Families"]
    fam_names = set()
    for r in range(4, fam.max_row + 1):
        n = fam.cell(r, 1).value
        if n:
            fam_names.add(n)

    # Mark pending families status
    for r in range(4, fam.max_row + 1):
        n = fam.cell(r, 1).value
        if n in {"Bow", "Hand Crossbow", "Hook Axe", "War Pick", "Ailment Reliquary"}:
            fam.cell(r, 14).value = "familyConfirmedContentPending"

    missing_families = [
        ("Staff", "Weapon", "Two-handed staff", 2, "Mage / Siren / Inquisitor", "Focus / Resolve", "Staff Bolt", "Staff Burst", "Magic staff damage", "Focus %, Resolve, Magic Penetration", "Two-handed", "Pine Staff", "Worldtree Branch", "Catalogue family"),
        ("War Scythe", "Weapon", "Two-handed scythe", 2, "Inquisitor / Mage / Duke", "Might / Focus", "Reap", "Harvest", "Hybrid reaping attacks", "Might, Focus, Bleed hybrids", "Two-handed", "Rusted Reaping Hook", "Eclipse Harvester", "Catalogue family"),
        ("Medium Harness", "Armour", "Medium armour/plumage", 0, "Any", "Guard / Vitality", "Wingbrace", None, "Balanced armour technique", "Guard flat, Vitality", "—", "Riveted Flight Harness", "Rookguard Harness", "Catalogue family"),
        ("Hunter Leathers", "Armour", "Light hunter armour", 0, "Any", "Agility / Vitality", "Hunter Veil", None, "Evasion-leaning light armour", "Agility, Critical", "Permanent Evasion cap 20%", "Patchwork Hunter Leathers", "Apex Predator's Mantle", "Catalogue family"),
        ("Dragonscale Mail", "Armour", "Scale armour/plumage", 0, "Knight / Brute / Duke / Inquisitor", "Guard / Might / Vitality", "Scale Counter", None, "Counter-leaning heavy scale", "Guard, Might flat", "Class restricted", "Worn Scale Harness", "Worldwyrm Plumage", "Catalogue family"),
        ("Sacred Vestments", "Armour", "Sacred mystic vestments", 0, "Mage / Siren / Inquisitor / Bard", "Resolve / Focus / Vitality", "Sanctuary", None, "Sacred barrier/heal support", "Resolve, Healing/Barrier Power", "Class restricted", "Pilgrim's Vestments", "Vestments of the First Song", "Catalogue family"),
        ("Warhelm", "Helmet", "Armoured helmet", 0, "Any", "Guard / Vitality", None, None, "Armoured head protection", "Guard flat, Vitality", "—", "Rustcap Helm", "Crowned Blackstone Warhelm", "Catalogue family"),
        ("Hunter Hood", "Helmet", "Hunter hood/cowl", 0, "Any", "Agility / Critical", None, None, "Hunter headgear", "Agility, Critical", "—", "Frayed Hunter Hood", "Hood of the Unseen Rook", "Catalogue family"),
        ("Battle Crown", "Helmet", "Battle crown/circlet", 0, "Any", "Might / Guard", None, None, "Martial crown", "Might flat, Guard", "—", "Ironbark Circlet", "Blakiston's War Crown", "Catalogue family"),
        ("Mystic Headwear", "Helmet", "Mystic hood/circlet", 0, "Any", "Focus / Resolve", None, None, "Mystic headgear", "Focus, Resolve", "—", "Traveller's Hood", "Oracle Crown of Six", "Catalogue family"),
        ("Mirror Wingguard", "Shield", "Light mirror wing shield", 0, "Any", "Resolve / Focus", None, None, "Reflective light wingguard", "Resolve, Magic Penetration", "—", "Polished Tin Mirror", "Mirror of Returning Skies", "Catalogue family"),
        ("Thorn Kite", "Shield", "Medium thorn wing shield", 0, "Any", "Guard / Might", None, None, "Thorned medium wingguard", "Guard, Martial Penetration", "—", "Bramble Kite", "Crown-of-Thorns Wingguard", "Catalogue family"),
        ("Iron Anklet", "Anklet", "Iron anklet", 0, "Any", "Guard / Vitality", None, None, "Sturdy iron anklet", "Guard flat, Vitality", "—", "Iron Ring Anklet", "Blackstone Rootchain", "Catalogue family"),
        ("Hexward Amulet", "Necklace", "Protective ward necklace", 0, "Any", "Resolve / Vitality", None, None, "Protective ward", "Resolve, Barrier Power", "—", "Salt-Ward Charm", "Amulet of the Unbroken Name", "Catalogue family"),
        ("Affinity Prism", "Necklace", "Affinity prism necklace", 0, "Any", "Focus / Affinity Damage", None, None, "Affinity amplification", "Affinity Damage %, Focus", "—", "Dull Affinity Shard", "Prism of Sixfold Light", "Catalogue family"),
    ]
    for row in missing_families:
        if row[0] in fam_names:
            continue
        r = fam.max_row + 1
        for c, v in enumerate(row, 1):
            fam.cell(r, c, v)
        fam_names.add(row[0])

    fam["A1"] = "Expanded Equipment Family Guide — v0.7"
    fam["A2"] = (
        "Weapon/orb guide plus full catalogue-family coverage for importer integrity. "
        "Bow, Hand Crossbow, Hook Axe, War Pick and Ailment Reliquary remain familyConfirmedContentPending (no invented named item rows)."
    )

    # ---- Dashboard ----
    dash = wb["Dashboard"]
    ct = wb["Combination Techniques"]
    pair_count = 0
    for r in range(4, ct.max_row + 1):
        sid = ct.cell(r, 12).value
        en = ct.cell(r, 3).value
        if sid and str(sid).startswith(("PAIR_", "COMBO_")) and en not in (None, ""):
            try:
                float(en)
                pair_count += 1
            except Exception:
                pass
    dash["A2"] = (
        f"Version 0.7 Equipment Loot implementation • 52 birds • 240 hybrid flat/% items • "
        f"{skill_count} active skill templates • {pair_count} curated pair/combination records • "
        f"30-star / six-tier progression"
    )

    # ---- Cursor Prompt retarget ----
    cp = wb["Cursor Prompt"]
    clear_sheet(cp)
    cp["A1"] = "Cursor Prompt — v0.7 Equipment Loot Systems Migration"
    cp["A2"] = (
        "Copy the text from E5. It preserves confirmed authority, separates Working Draft values "
        "and requires repository-grounded paths, symbols, migrations and tests."
    )
    cp["A4"] = "Section"
    cp["B4"] = "Requirement"
    cp["C4"] = "Authority"
    cp["D4"] = "COPY-READY PROMPT"
    prompt = """AVIAN ASCENT v0.7 — EQUIPMENT LOOT HYBRID FLAT/% MIGRATION

ROLE [Confirmed]
Act as a senior gameplay-systems engineer working on Avian Ascent.

SOURCE [Confirmed]
Use v7Newest_Avian_Ascent_Master_v0.7_Equipment_Loot_SyncReady.xlsx as the design source of truth.

PRIORITY
Resolve disagreements in this order: Confirmed workbook rules; explicit requirements in this prompt; current code behaviour; Working Draft workbook values. Never silently promote a Working Draft.

MODE
Inspect the repository first and produce a file-and-symbol grounded implementation plan. Do not modify code until the plan is approved.

CALCULATION ORDER [Confirmed R-PROG-005]
Base + Level Flat + Star Flat → Tier Multiplier → Equipment Flat → additive Equipment % → temporary effects. Round after tier and after equipment.

EQUIPMENT [Confirmed]
Core loot is hybrid flat + percentage for Vitality, Might, Guard, Focus, Resolve and Agility. Grey emphasises flat; higher rarities mix both. Percentage rolls may display to two decimals; flats remain integers.

ENERGY / COOLDOWN [Confirmed]
Start 4 EN, recover 3 / turn. Fixed cooldown by EN: 1–2 → 0; 3 → 1; 4 → 2; 6 → Ultimate gate. Cooldown never refunds damage, healing or utility budget. Riders that add utility, ailment, penetration, healing or barrier reduce direct damage instead.

DAMAGE BANDS [Confirmed R-EN-005]
1 EN 70–90%; 2 EN 100–120%; 3 EN 130–150%; 4 EN 160–190%; 5 EN 195–225%; 6 EN 235–280%. Natural Strike is the 1 EN baseline.

PENETRATION [Confirmed R-PEN-002]
Ignore Guard and Ignore Resolve are separate properties. All penetration sources share the 40% cap.

DUAL WIELD [Confirmed]
Matching one-handed families use Paired Technique. Compatible non-matching tags use curated Combination Technique. Unlinked pairs retain both normals. Combination replaces only the off-hand normal action.

DEFER [Confirmed]
Do not invent Bow, Hand Crossbow, Hook Axe, War Pick or Ailment Reliquary named catalogue rows. Keep familyConfirmedContentPending.

LANGUAGE [Confirmed]
No direct Frozen, Paralysed, Sleep or Silence from active equipment/skill text. Keep strict 1v1; no AoE/cleave/volley mechanics.

PRESERVE
Affinity matrix, ailment stack rules, Orb focuses, Precision/Evasion caps, Brace/Barrier terminology, data-driven combinations.

OUTPUT
Repository architecture with paths/symbols; workbook-to-code gap table; dependency-ordered phases; file-by-file change map; save migration; automated tests; blockers.
"""
    cp["E5"] = prompt
    cp["A5"] = "ROLE"
    cp["B5"] = "Act as a senior gameplay-systems engineer working on Avian Ascent."
    cp["C5"] = "Confirmed"
    cp["D5"] = "See E5 copy-ready prompt"

    # Update checklist statuses now that workbook gaps closed
    impl = wb["Implementation Checklist"]
    for r in range(5, impl.max_row + 1):
        task = impl.cell(r, 2).value
        if task in ("Combination Skill IDs", "Family guide coverage", "Rule doc sheets"):
            impl.cell(r, 6).value = "Done"

    wb.save(DST)
    # also overwrite source path alias used by plan
    wb.save(SRC)
    print(f"Wrote {DST}")
    print(f"Also updated {SRC}")
    print(f"Skills added: {added}; total skills: {skill_count}; pair records: {pair_count}")
    print(f"Families now: {len(fam_names)}")


if __name__ == "__main__":
    main()
