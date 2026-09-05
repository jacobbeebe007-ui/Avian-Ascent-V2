#!/usr/bin/env python3
"""Merge Precision, Combat v2.1 and Structured Effects master into one workbook."""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DUMP = Path("/tmp/wb-dump")
OUT = ROOT / "Avian_Ascent_Current_Master_v2.1.xlsx"

HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F4A3C")
LOCK_FILL = PatternFill("solid", fgColor="C8E6C9")
WARN_FILL = PatternFill("solid", fgColor="FFF3CD")
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")


def tsv_rows(kind: str, name: str) -> list[list[str]]:
    path = DUMP / kind / f"{name}.tsv"
    if not path.exists():
        return []
    rows = []
    with path.open(newline="") as f:
        for line in csv.reader(f, delimiter="\t"):
            rows.append(line)
    return rows


def style_header(ws, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(1, col)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws, widths: list[int] | None = None) -> None:
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 10
        for cell in col[:40]:
            longest = max(longest, min(48, len(str(cell.value or ""))))
        ws.column_dimensions[letter].width = longest + 2


def write_rows(ws, rows: list[list], widths: list[int] | None = None, header=True) -> None:
    for r in rows:
        ws.append(list(r))
    if header and rows:
        style_header(ws, max(len(r) for r in rows))
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP
    autosize(ws, widths)


def copy_tsv(wb: Workbook, title: str, kind: str, name: str, widths: list[int] | None = None) -> None:
    ws = wb.create_sheet(title[:31])
    rows = tsv_rows(kind, name)
    write_rows(ws, rows, widths, header=True)


def main() -> None:
    wb = Workbook()

    # --- Dashboard ---
    ws = wb.active
    ws.title = "Dashboard"
    write_rows(ws, [
        ["Avian Ascent — Current Master v2.1", "", "", ""],
        ["Merged from Combat Workbook v2.1, Bird Precision System, and Current Master v1.6 Structured Effects.", "", "", ""],
        ["Area", "Current rule", "Primary sheet", "State"],
        ["Authoritative combat design", "v2.1 locked rules; Attack Power / Health are next-foundation (runtime still weapon-first + Vitality×3).", "V2 Core Rules; Current Rules", "LOCKED"],
        ["Cooldown-free AP budget", "Equal EV/AP: 1=0.45, 2=1.00, 3=1.50, 4=2.10. Ordinary CDs = 0 once Attack Power is live.", "AP & Skill Budget; AP Expectation Tests", "LOCKED"],
        ["Hybrid damage", "Portions chip matching pools. Health gate = total − mean(start Armour, start Magic Armour).", "Hybrid Damage", "LIVE"],
        ["Affinity", "Dominant ×1.10 / Neutral ×1.00 / Resisted ×0.90 (22% relative, was 50%).", "Affinity", "LIVE"],
        ["Ultimate Meter", "6 × AP, once per landed action, not per hit. Utility 0. Cap 24/turn. Ultimate 6 AP + 100 meter.", "Ultimate Meter", "LIVE"],
        ["Sequential encounters", "Health, AP, buffs, player ailments and meter persist. Protection refills to normal max. Fortify/Ward overflow expires.", "Sequential Encounters", "LIVE"],
        ["Species rarity", "Grey and Orange birds are equally viable. Rarity is identity and item access, not a species damage multiplier.", "Species Rarity", "LOCKED"],
        ["Precision", "Class + Size + Species. Runtime field acc. 52/52 audited.", "Bird Stats; Class Precision Reference", "CURRENT"],
        ["Structured effects", "105 skills / 170 effect rows remain the skill-mechanic source until weapon republish.", "Skills; Skill Effects", "CURRENT"],
        ["Telemetry", "≥200 runs/matchup. Required: actionChosen, unusedAP, Fortify, Ward, ailment attempt/ok/gated.", "Telemetry Spec", "REQUIRED"],
        ["Next runtime phase", "Replace Attack Power and Health formulas, then clear ordinary cooldowns and retune skill numbers.", "Implementation Plan", "REQUIRED"],
    ], [28, 88, 40, 12])

    # --- Investigation Register ---
    ws = wb.create_sheet("Investigation Register")
    write_rows(ws, [
        ["Priority", "System", "Main risk", "Decision", "State"],
        ["1", "Cooldown-free action selection", "One repeatable 3 AP skill becomes the only correct action.", "Flatten 3 AP from 1.55 → 1.50 (0.50/AP, same as 2 AP). 1 AP stays 0.45 (flexibility tax). 4 AP 2.10 (packed-damage premium vs pools). Tactical purpose, not raw EV, differentiates actions.", "LOCKED"],
        ["2", "Full progression curve", "Damage, Health and protection must stay balanced at 1 / 5 / 10 / 15 / 20 / 25.", "Attack Power = Weapon + 2×stat. Health = Size 125–140 + 5×VIT + 5×(L−1). See Progression Curve. Runtime still uses weapon-first and Vitality×3 until Phase 1.", "PROTOTYPE"],
        ["3", "Armour and Magic Armour", "Oversized pools recreate 10-turn Mage/Siren stalls.", "Neither pool may exceed 55% of expected 3-turn incoming. Restoration ≤35% of one 3 AP hit. Fortify/Ward ≤75% of one expected enemy turn. Anti-stall flags 3 rounds of zero progress.", "LOCKED"],
        ["4", "Hybrid damage", "50/50 hybrids must empty both pools before Health.", "Mean-pool Health gate. Equal 20/20 pools: 40 hybrid deals 20 Health (same as a specialist). Mismatched pools sit between the two specialists. Live in protection-pools.js.", "LIVE"],
        ["5", "All 52 bird identities", "Class, stats, passive, utility and loadout need a species pass.", "Foundation recalculated in Bird Recalibration. Barn Owl → Rogue (locked). Passives/utilities/loadouts remain REQUIRED next pass.", "PARTIAL"],
        ["6", "Equipment contribution", "Gear overpowers bird stats and erases levels.", "Weapon is the only Attack Power source from items. Whole-loadout flat-stat budgets Grey 6 … Orange 24. Rarity does not multiply the action coefficient.", "LOCKED"],
        ["7", "Ultimate Meter", "Cheap and multi-hit attacks fill the meter too fast.", "Award 6 × AP once per landed action. Multi-hit = one award. Utility 0. Turn cap 24 so 4×1 AP = one 4 AP action.", "LIVE"],
        ["8", "Status availability", "Starters never apply ailments after breaking protection.", "Every starter kit must grant Technique B with a Health-gated ailment and a Grey defence action. See Status Availability.", "REQUIRED"],
        ["9", "Enemy AI", "AI wastes AP, repeats buffs, ignores protection breaks.", "Priority: lethal → break+follow-up → defence if survival improves → status if it can resolve → EV within 5%. No repeat buff while a stronger same-direction effect is active. Story may leave 1 AP; Standard unused AP ≤ 0.5/turn.", "LOCKED"],
        ["10", "Sequential encounters", "Unclear carry of Health, pools, AP, buffs, ailments, meter.", "Health / AP / buffs / player ailments / meter persist. Protection refills to normal max. Fortify/Ward overflow expires. Once-per-battle flags persist through the sequence.", "LIVE"],
        ["11", "Affinity balance", "1.20 vs 0.80 is a 50% relative damage swing.", "1.10 / 1.00 / 0.90 (22% relative). Chart relationships unchanged.", "LIVE"],
        ["12", "Species rarity", "Grey vs Orange permanent power.", "No bird-rarity damage multiplier. Grey and Orange are equally viable at the same gear milestone. 30/32/34 attribute points stay as authored identity only.", "LOCKED"],
    ], [10, 28, 44, 88, 12])

    # --- AP tests ---
    ws = wb.create_sheet("AP Expectation Tests")
    write_rows(ws, [
        ["AP EXPECTATION LAB — Attack Power 21.5 (Weapon 3.5 + 2×9). Defence ignored for EV; packed vs split still matters vs pools.", "", "", "", "", ""],
        ["Test", "Spend", "Actions", "Raw damage", "Damage/AP", "Reading"],
        ["Four 1 AP vs one 4 AP", "4", "4×0.45 vs 2.10", "38.70 vs 45.15", "0.45 vs 0.538", "4 AP wins raw and vs protection (one packet overflows). 1 AP is leftover/setup."],
        ["Two 2 AP vs one 4 AP", "4", "2×1.00 vs 2.10", "43.00 vs 45.15", "0.50 vs 0.538", "4 AP modestly better. Banking is rewarded, not mandatory."],
        ["Repeat 3 AP every turn", "3/turn", "1.50 each turn from 4→1 leftover→4", "32.25/turn", "0.50", "Equal EV/AP to 2 AP. Dominates only if it also carries a free ailment — then drop its coeff to 1.35–1.40."],
        ["Alternate 2 AP and 4 AP", "6 / 2 turns", "1.00 + 2.10", "33.325/turn", "0.517", "Slightly above 3 AP spam. Saved AP has a purpose."],
        ["Buff once, attack three turns", "2+2 / 3 / 2+1", "Major +2 stat for 3 turns", "112.7 vs 96.75 no-buff", "—", "Setup pays ~16% over three turns without becoming a per-turn tax."],
        ["Rogue 1 AP refund + 1 AP actions", "4 + 1 refund", "2 AP then 1+1", "1.00+0.45+0.45=1.90 vs 1.50", "—", "Allowed once/turn, first-actor, non-Basic, no chain. Opening burst is the perk, not a loop."],
        ["Multi-hit Ultimate Meter", "same AP", "4×1 vs 1×4", "meter 24 vs 24", "6/AP", "Once per landed action. Hits do not multiply meter."],
        ["Fortify or Ward every second turn", "4 then 3", "4 AP defence, then ≤3 AP", "cannot repeat on 3 AP", "—", "No cooldown required. Amount ≤75% of one expected enemy turn."],
        ["Old 3 AP coeff 1.55", "3/turn", "1.55", "33.325 vs 32.25", "0.517 vs 0.50", "REJECTED — 3.4% EV edge plus ailment rider made 3 AP the only correct button."],
    ], [32, 16, 36, 22, 14, 72])

    # --- Hybrid ---
    ws = wb.create_sheet("Hybrid Damage")
    write_rows(ws, [
        ["V2-HYB-001  Mean-pool Health gate", "", "", ""],
        ["Rule", "Physical portion → Armour. Magic portion → Magic Armour. Health = max(portion leftovers, total − mean(start Armour, start Magic Armour)). Extra Health is a rebate: over-absorbed pool damage is restored so applied damage never exceeds the incoming packet.", "", ""],
        ["Why", "A 50/50 hybrid no longer has to empty both pools before dealing Health. Specialists still win into the matching empty pool. Hybrids sit between the two specialists into mismatched pools.", "", ""],
        ["Case", "Incoming", "Pools", "Health / reading"],
        ["Specialist physical 40", "40 physical", "20 / 20", "20 Health, Armour 0. Reference."],
        ["Hybrid 20+20 (old rule)", "20 + 20", "20 / 20", "0 Health. FAIL — this is why Bard telemetry was 6%."],
        ["Hybrid 20+20 (adopted)", "20 + 20", "20 / 20", "20 Health, both pools 10. Parity + chips both."],
        ["Hybrid 20+20", "20 + 20", "10 / 30", "20 Health (mean gate 20). Midway between physical 30 and magic 10."],
        ["Hybrid 20+20", "20 + 20", "0 / 40", "20 Health. Midway between physical 40 and magic 0."],
        ["Hybrid 20+20", "20 + 20", "40 / 40", "0 Health. Same as a specialist. Thick dual tanks still stall — that is an armour-budget problem."],
        ["Coefficient", "Hybrid actions use the same AP coefficients as pure damage. Do not apply a second coeff tax (Bard Song 70% at 2 AP was a double tax).", "", ""],
        ["Runtime", "Avian.protection.applyHybridDamageThroughProtection — live.", "", ""],
    ], [28, 22, 16, 72])

    # --- Affinity ---
    ws = wb.create_sheet("Affinity")
    write_rows(ws, [
        ["Relationship", "Old", "Adopted", "Relative vs opposite", "State"],
        ["Dominant", "1.20", "1.10", "1.10 / 0.90 = 1.222 (22%)", "LIVE"],
        ["Neutral / same", "1.00", "1.00", "—", "LIVE"],
        ["Resisted", "0.80", "0.90", "was 1.20/0.80 = 1.50 (50%)", "LIVE"],
        ["Chart", "Unchanged six-way wheel (Earth/Sky/Storm/Day/Night/Water).", "", "", "LOCKED"],
        ["Do not", "Stack a second affinity multiplier on the action coefficient or on weapon rarity.", "", "", "LOCKED"],
    ], [20, 12, 12, 36, 10])

    # --- Ultimate Meter ---
    ws = wb.create_sheet("Ultimate Meter")
    write_rows(ws, [
        ["AP", "Award", "4× this action", "Equals"],
        ["1", "6", "24", "one 4 AP action"],
        ["2", "12", "24 (two actions)", "one 4 AP action"],
        ["3", "18", "—", "3+1 = 24"],
        ["4", "24", "—", "turn cap"],
        ["5", "24 (capped)", "—", "turn cap 24"],
        ["6 / Ultimate", "0", "—", "spends the meter"],
        ["Utility", "0", "—", "no meter from Fortify/Ward/songs without a landed hit"],
        ["Multi-hit", "Once per landed action", "—", "four hits still one award"],
        ["Miss", "0", "—", "AP is still spent"],
        ["Rule", "Meter gained only by landed direct hits. Ultimate costs 6 AP and a full 100 meter.", "", ""],
    ], [18, 22, 22, 28])

    # --- Sequential ---
    ws = wb.create_sheet("Sequential Encounters")
    write_rows(ws, [
        ["Resource", "Carry into next enemy", "Why"],
        ["Health", "Persist (no refill)", "The sequence is one exertion."],
        ["Armour / Magic Armour", "Refill to normal maximum", "New foe; do not inherit a broken wall from the last kill."],
        ["Fortify / Ward overflow", "Expire at cleanup", "Temporary max does not walk into the next fight."],
        ["AP", "Persist, cap 6", "Banking is a real decision."],
        ["Buffs / debuffs on player", "Persist remaining duration", "Setup can pay across the sequence."],
        ["Ailments on player", "Persist", "Poisoned stays poisoned."],
        ["Ailments on defeated enemy", "Discard", "The next bird is a new combatant."],
        ["Ultimate Meter", "Persist", "Meter is a sequence resource, not a single-duel resource."],
        ["Once-per-battle flags", "Persist through the sequence", "A sequence is one battle for perk clocks."],
        ["Enemy AP / meter / statuses", "Fresh", "Each foe starts at 4 AP, 0 meter, full pools."],
    ], [28, 32, 56])

    # --- Species rarity ---
    ws = wb.create_sheet("Species Rarity")
    write_rows(ws, [
        ["Question", "Decision", "State"],
        ["Are Grey and Orange birds equally viable?", "Yes, at the same equipment milestone.", "LOCKED"],
        ["Does rarity grant permanent damage?", "No. Remove any Grey=1.00 … Orange=1.20 bird combat multiplier.", "LOCKED"],
        ["What may differ?", "Authored attribute budget (30 / 32 / 34), stat caps, passives, and which items they can find.", "LOCKED"],
        ["Where does power come from?", "Level, class growth, and equipment rarity — not the bird's colour band.", "LOCKED"],
        ["Orange uniqueness", "Per-run items and encounter design, not a hidden 20% species tax.", "LOCKED"],
    ], [40, 72, 10])

    # --- Progression curve ---
    ws = wb.create_sheet("Progression Curve")
    rows = [
        ["Level", "Knight Might", "Grey weapon AP", "Orange weapon AP", "Tiny HP (VIT 0)", "Medium HP (VIT 5)", "Giant HP (VIT 9)", "Grey heavy prot", "Orange heavy prot", "Reading"],
        ["1", "8", "3+16=19", "9+16=25", "125", "156", "185", "36", "88", "Weapon is visible; size gap 48% including VIT, 12% size-only."],
        ["5", "10", "3+20=23", "9+20=29", "145", "176", "205", "36", "88", "+4 Attack Power and +20 HP from four levels."],
        ["10", "12", "3+24=27", "9+24=33", "170", "201", "230", "36", "88", "Level now outruns a Grey weapon."],
        ["15", "14", "3+28=31", "9+28=37", "195", "226", "255", "36", "88", "Protection budgets must be retuned with rarity, not level."],
        ["20", "16", "3+32=35", "9+32=41", "220", "251", "280", "36", "88", "Same-rarity mirrors stay in the 5–7 round band if incoming 3 AP ≈ 0.5 of unprotected HP / 6."],
        ["25", "18", "3+36=39", "9+36=45", "245", "276", "305", "36", "88", "Orange weapon is +6 AP over Grey — gear matters, levels still matter."],
        ["Guardrail", "Highest÷lowest unprotected L1 Health ≤ 1.50", "Total durability ≤ 1.65 before tank gear", "", "", "", "", "", "", "From Size & Health."],
    ]
    write_rows(ws, rows, [10, 14, 16, 16, 16, 18, 16, 16, 18, 64])

    # --- Telemetry spec ---
    ws = wb.create_sheet("Telemetry Spec")
    write_rows(ws, [
        ["Field", "Why", "Grouping", "Warning"],
        ["playerWin", "Difficulty outcome", "class / tier / loadout / size", "Story <50% or >90%; Standard mirror outside 45–55%"],
        ["rounds", "Pacing", "mode / class / tier", "Median outside 5–7 (Story 4–7)"],
        ["unusedAP", "AI / player efficiency", "actor / reason", "Standard average >0.5 / turn"],
        ["actionChosen", "Skill distribution", "skill / source / AP", "Any equipped core action never selected"],
        ["fortify / ward", "Defence reachability", "source / tier", "Zero in an eligible pack"],
        ["ailmentAttempt / ok / gated", "Status reachability", "ailment / pool / round", "ok=0 in an eligible starter pack"],
        ["damageSource", "Progression attribution", "weapon / stat / coeff / crit / affinity", "Cannot reconcile final damage"],
        ["effectiveDurability", "Size / loadout", "bird / size / tier", "Same-tier ratio >1.65"],
        ["firstActorWin", "Initiative", "matchup / order", ">55% Standard after paired order"],
        ["Runs", "≥200 per matchup", "fixed seeds + random seeds", "Old 50-run Story pack is not valid for v2.1"],
        ["Do not start", "Until Attack Power, Health, single hit roll, no ordinary CDs, and AI defence/status use are live.", "", "Old telemetry cannot validate v2.1"],
    ], [28, 36, 28, 48])

    # --- Status availability ---
    ws = wb.create_sheet("Status Availability")
    write_rows(ws, [
        ["Starter class", "Required attack route", "Required defence / status route", "Gate"],
        ["Knight", "Grey Beak Hammer or Lance Technique that can Fracture on Health", "Grey Shield Fortify or Heavy Plumage Brace", "Armour → 0 then Health"],
        ["Rogue", "Grey Talon Blade / Dagger Pinion Technique B Bleed on Health", "Grey Light Plumage Featherstep", "Armour → 0 then Health"],
        ["Mage", "Grey Wand / Hexwood source ailment on Health", "Grey Focus Orb or Wand Arcane Ward", "Magic Armour → 0 then Health"],
        ["Siren", "Grey Lament / Hexwood Resolve Down or source ailment on Health", "Grey Ward from Orb or Plumage", "Magic Armour → 0 then Health"],
        ["Inquisitor", "Grey War Scythe Health-damage lifesteal (counts as Health route)", "Grey Plumage restore", "Matching pool then Health"],
        ["Bard", "Grey Bard Song hybrid that can land Health via the mean-pool gate", "Grey restore / Chorus protection", "Either pool then Health"],
        ["Brute", "Grey Beak Hammer Dazed on Health", "Grey Heavy Plumage Brace or Fortifying Slam", "Armour → 0 then Health"],
        ["Failure", "If Fortify, Ward or ailments.ok stay 0 in an eligible starter pack, the kit is incomplete — do not retune coefficients yet.", "", ""],
    ], [16, 52, 44, 28])

    # --- Change log ---
    ws = wb.create_sheet("Change Log")
    write_rows(ws, [
        ["Version", "Change"],
        ["v2.1 master", "Merged Combat Workbook v2.1, Bird Precision System, and Current Master v1.6 Structured Effects."],
        ["v2.1 master", "Locked AP coefficients 0.45 / 1.00 / 1.50 / 2.10 after EV tests. Rejected 3 AP at 1.55."],
        ["v2.1 master", "Adopted hybrid mean-pool Health gate. Implemented in runtime."],
        ["v2.1 master", "Affinity 1.10 / 0.90. Ultimate Meter 6×AP with 24/turn cap. Sequential carry live."],
        ["v2.1 master", "Species rarity is not a damage multiplier. Grey and Orange equally viable."],
        ["v2.1 master", "Runtime still uses weapon-first damage and Vitality×3 Health until Phase 1."],
        ["Sources", "Avian_Ascent_Combat_Workbookv2.1.xlsx · Avian_Ascent_Bird_Precision_System.xlsx · Avian_Ascent_Current_Master_v1.6_Structured_Effects Updated.xlsm"],
    ], [16, 100])

    # Combat v2.1 design sheets (updated copies)
    copy_tsv(wb, "V2 Core Rules", "combat", "V2_Core_Rules")
    copy_tsv(wb, "AP & Skill Budget", "combat", "AP_Skill_Budget")
    copy_tsv(wb, "Damage & Progression", "combat", "Damage_Progression")
    copy_tsv(wb, "Buff & Duration", "combat", "Buff_Duration")
    copy_tsv(wb, "Size & Health", "combat", "Size_Health")
    copy_tsv(wb, "Class Framework", "combat", "Class_Framework")
    copy_tsv(wb, "Bird Recalibration", "combat", "Bird_Recalibration")
    copy_tsv(wb, "Weapon Framework", "combat", "Weapon_Framework")
    copy_tsv(wb, "Weapon Distribution", "combat", "Weapon_Distribution")
    copy_tsv(wb, "Equipment Rework", "combat", "Equipment_Rework")
    copy_tsv(wb, "Defence & Status", "combat", "Defence_Status")
    copy_tsv(wb, "Enemy & AI", "combat", "Enemy_AI")
    copy_tsv(wb, "Balance Lab", "combat", "Balance_Lab")
    copy_tsv(wb, "Implementation Plan", "combat", "Implementation_Plan")
    copy_tsv(wb, "Telemetry Diagnosis", "combat", "Telemetry_Diagnosis")

    # Patch AP budget sheet with the locked 1.50 / 2.10 row if present
    ap = wb["AP & Skill Budget"]
    for row in ap.iter_rows(min_row=1, max_row=12, max_col=3):
        if str(row[0].value) == "3":
            row[1].value = 1.50
            row[2].value = 0.50
            row[1].fill = LOCK_FILL
        if str(row[0].value) == "4":
            row[1].value = 2.10
            row[2].value = 0.525
            row[1].fill = LOCK_FILL
        if str(row[0].value) == "5":
            row[1].value = 2.70
            row[2].value = 0.54
        if str(row[0].value) == "6":
            row[1].value = 3.30
            row[2].value = 0.55

    # Precision
    copy_tsv(wb, "Class Precision Reference", "precision", "Class_Precision_Reference")
    copy_tsv(wb, "Size Precision Reference", "precision", "Size_Precision_Reference")
    copy_tsv(wb, "Species Precision Audit", "precision", "Species_Precision_Audit")
    copy_tsv(wb, "Loadout Builder", "precision", "Loadout_Builder")
    copy_tsv(wb, "Weapon Families Precision", "precision", "Weapon_Families")

    # Master runtime-authoritative sheets (keep +3 HP so bird-stat verify still matches live data)
    copy_tsv(wb, "Current Rules", "master", "Current_Rules")
    copy_tsv(wb, "Core Rules", "master", "Core_Rules")
    copy_tsv(wb, "Bird Stats", "master", "Bird_Stats")
    copy_tsv(wb, "Bird Abilities", "master", "Bird_Abilities")
    copy_tsv(wb, "Class Perks", "master", "Class_Perks")
    copy_tsv(wb, "Passives & Perks", "master", "Passives_Perks")
    copy_tsv(wb, "Affinities & Ailments", "master", "Affinities_Ailments")
    copy_tsv(wb, "Fortify & Ward", "master", "Fortify_Ward")
    copy_tsv(wb, "Skills", "master", "Skills")
    copy_tsv(wb, "Skill Effects", "master", "Skill_Effects")
    copy_tsv(wb, "Effect Definitions", "master", "Effect_Definitions")
    copy_tsv(wb, "Skill Rules", "master", "Skill_Rules")
    copy_tsv(wb, "Weapon Skill Library", "master", "Weapon_Skill_Library")
    copy_tsv(wb, "Equipment Skill Library", "master", "Equipment_Skill_Library")
    copy_tsv(wb, "Stat Definitions", "master", "Stat_Definitions")
    copy_tsv(wb, "Effect Tiers", "master", "Effect_Tiers")
    copy_tsv(wb, "Scaling Model", "master", "Scaling_Model")
    copy_tsv(wb, "Progression Rules", "master", "Progression_Rules")
    copy_tsv(wb, "Bird & Classes", "master", "Bird_Classes")

    # Annotate Current Rules
    cr = wb["Current Rules"]
    cr.insert_rows(2)
    cr["A2"] = (
        "Runtime still uses these v1.6 conversions (Vitality × 3, weapon-first 2.5%). "
        "Adopted next-foundation rules are on V2 Core Rules, AP Expectation Tests, Hybrid Damage, Affinity and Ultimate Meter."
    )
    cr["A2"].fill = WARN_FILL
    cr["A2"].alignment = WRAP

    wb.save(OUT)
    print("wrote", OUT, "sheets", len(wb.sheetnames))


if __name__ == "__main__":
    main()
